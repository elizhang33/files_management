import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
import re
from collections import defaultdict, Counter
import logging
import datetime
import shutil

# Initialize logging
logging.basicConfig(filename='csvtoxlsx_count.log', level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')  # <-- Configure the logger with time

LOCK_FILE = "XXX"
FLOORSETS_DIR = 'XXX'
CC_COUNT_FILE = 'XXX'
ARTICLE_PATTERN = re.compile(r'([^_]+)')
DATE_PATTERN = re.compile(r'(\d{2}_\d{2}_\d{4})')
NON_WORD_PATTERN = re.compile(r'\W+')

def remove_unwanted_substrings(string):
    pattern = re.compile(r'2023_.*$')
    return pattern.sub('2023', string)

def format_filename(filename):
    '''Replaces all non-alphanumeric characters with underscores, PRIORITY with empty string'''
    filename = re.sub(r'(?i)PRIORITY', '', filename)
    return re.sub(r'\W', '_', filename)


def convert_csv_to_excel(csv_dir, excel_dir):
    # Create the Excel directory if it doesn't exist
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)

    # Get the list of CSV files in the CSV directory
    csv_files = [file for file in os.listdir(csv_dir) if file.endswith(".csv")]

    for csv_file in csv_files:
        csv_path = os.path.join(csv_dir, csv_file)

        # Read the CSV file into a pandas DataFrame
        df = pd.read_csv(csv_path)

        # Create the Excel file path
        excel_file = os.path.splitext(csv_file)[0] + ".xlsx"
        excel_path = os.path.join(excel_dir, excel_file)

        # Write the DataFrame to the Excel file
        df.to_excel(excel_path, index=False)

        # Autofit the columns in the Excel sheet
        autofit_columns(excel_path)

        # Add table to the Excel sheet
        add_table(excel_path, df)


def autofit_columns(excel_path):
    # Load the Excel workbook
    workbook = load_workbook(excel_path)

    # Get the active sheet
    sheet = workbook.active

    # Autofit columns
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length

    # Save the modified workbook
    workbook.save(excel_path)


def add_table(excel_path, df):
    # Load the Excel workbook
    workbook = load_workbook(excel_path)

    # Get the active sheet
    sheet = workbook.active

    # Convert DataFrame to rows
    rows = dataframe_to_rows(df, index=False, header=True)

    # Determine the range of the table
    start_cell = sheet.cell(row=1, column=1)
    end_cell = sheet.cell(row=df.shape[0]+1, column=df.shape[1])

    # Add the table to the sheet
    table = Table(displayName="Table1", ref=f"{start_cell.coordinate}:{end_cell.coordinate}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Save the modified workbook
    workbook.save(excel_path)

def check_lock():
    return os.path.exists(LOCK_FILE)

def acquire_lock():
    with open(LOCK_FILE, "w") as f:
        f.write("Locked")

def release_lock():
    os.remove(LOCK_FILE)

def update_floorset_and_count(csv_path, floorset):
    '''Update the floorset's csv file with new filenames and count unique articles'''
    log_msg = f"Updating {floorset}..."
    print(log_msg)
    logging.info(log_msg)

    floorset_filename  = format_filename(floorset)
    floorset_csv_path = os.path.join(FLOORSETS_DIR, f"{floorset_filename }.csv")
    cc_count_df = pd.read_csv(CC_COUNT_FILE, index_col='Floorset')

    new_articles = set()
    
    # Check if the floorset file already exists. If not, we'll create it later.
    if os.path.exists(floorset_csv_path) and os.path.getsize(floorset_csv_path) > 0:
        # If it exists, read the current unique articles from the floorset's csv
        current_floorset_df = pd.read_csv(floorset_csv_path)
        current_articles = set(current_floorset_df.iloc[:,0].apply(lambda x: ARTICLE_PATTERN.match(x).group(1)))
    else:
        current_articles = set()
           
    filenames = []

    # Initialize dictionaries to hold views for each article
    article_views = defaultdict(set)

    
    with open(csv_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)  # Skip the first row (header)
        for row in reader:
            filename = row[1]  # Get filename from the second column
            article = ARTICLE_PATTERN.match(filename).group(1)

            # Extract view part considering the 'swatch' case
            view_part = filename.split('_')[1].split('-')[0]
            if view_part != 'swatch':
                article_views[article].add(view_part)

            filenames.append(filename)
            new_articles.add(article)

    # Calculate the set of new unique articles
    added_articles = new_articles - current_articles

    # Update the floorset's csv file and the count only if there are new unique articles
    if added_articles:
        # Check if the floorset exists in the DataFrame's index
        if floorset not in cc_count_df.index:
            # If not, append a new row with the count
            cc_count_df.loc[floorset] = {'count': len(added_articles)}
        else:
            # If it exists, update the count as before
            cc_count_df.at[floorset, 'count'] += len(added_articles)
        
        total_articles = cc_count_df.at[floorset, 'count']
        cc_count_df.to_csv(CC_COUNT_FILE)
        log_msg = f"Added {len(added_articles)} new unique articles to {floorset}. Total unique articles: {total_articles}."
        print(log_msg)
        logging.info(log_msg)

        if 'XXX' in floorset:
            # Classify articles as debut or flow based on their views
            debut_articles = set()
            flow_articles = set()
            for article, views in article_views.items():
                if all(view in views for view in ['100', '110', '120', '130']):
                    debut_articles.add(article)
                elif '100' in views or '000' in views:
                    if all(view not in views for view in ['110', '120', '130']):
                        flow_articles.add(article)

            # Filter to only include new and unique articles
            new_unique_debut_articles = debut_articles.intersection(added_articles)
            new_unique_flow_articles = flow_articles.intersection(added_articles)

            debut_count = len(new_unique_debut_articles)
            flow_count = len(new_unique_flow_articles)

            log_msg = f"Added {debut_count} New DEBUT articles and {flow_count} New FLOW articles to {floorset}."
            print(log_msg)
            logging.info(log_msg)
        
        # Write the articles to the floorset's csv file
        with open(floorset_csv_path, 'a+', newline='') as floorset_csv:
            floorset_csv.seek(0, os.SEEK_END)
            pos = floorset_csv.tell() - 1
            if pos >= 0:
                floorset_csv.seek(pos)
                if floorset_csv.read(1) != '\n':
                    floorset_csv.write('\n')

            writer = csv.writer(floorset_csv)
            for article in new_articles:
                for file in filenames:
                    if article in file:
                        writer.writerow([file])

    else:
        log_msg = f"No new unique articles added to {floorset}."
        print(log_msg)
        logging.info(log_msg)


def move_outdated_files(base_dir):
    prefix_folders = {"XXX": "XXX", "XXX": "XXX", "XXX": "XXX"}
    today = datetime.date.today()

    # Loop over prefix and their corresponding folder.
    for prefix, folder in prefix_folders.items():
        folder_path = os.path.join(base_dir, folder)

        # Create the folder if it does not exist.
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Loop over files in the base_dir.
        for file in os.listdir(base_dir):
            if file.startswith(prefix) and file.endswith(".xlsx"):
                date_str = re.search(r'(\d{2}_\d{2}_\d{4})\.xlsx$', file).group(1)
                file_date = datetime.datetime.strptime(date_str, '%m_%d_%Y').date()

                # Compare the file date to today's date.
                if file_date < today:
                    file_path = os.path.join(base_dir, file)
                    destination_path = os.path.join(folder_path, file)
                    
                    # Move file to the corresponding folder.
                    shutil.move(file_path, destination_path)
                    log_msg = f"Moved outdated file: {file} to {destination_path}"
                    print(log_msg)
                    logging.info(log_msg)


def main():
    if check_lock():
        log_msg = "Script is currently in use by XXXeone else."
        print(log_msg)
        logging.info(log_msg)
        return

    acquire_lock()

    try:
        xlsx_dir = 'XXX'
        
        # Call the function to move outdated files.
        move_outdated_files(xlsx_dir)

        csv_dir = 'XXX'
        convert_csv_to_excel(csv_dir, xlsx_dir)  

        # Rest of the main logic...

    finally:
        release_lock()

if __name__ == "__main__":
    main()

