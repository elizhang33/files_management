import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

LOCK_FILE = "XXX"

def convert_csv_to_excel(csv_dir, excel_dir):
    # Create the Excel directory if it doesn't exist
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)

    # Get the list of CSV files in the CSV directory
    csv_files = [file for file in os.listdir(csv_dir) if file.endswith(".csv")]

    for csv_file in csv_files:
        csv_path = os.path.join(csv_dir, csv_file)

        # Read the CSV file into a pandas DataFrame
        df = pd.read_csv(csv_path)

        # Create the Excel file path
        excel_file = os.path.splitext(csv_file)[0] + ".xlsx"
        excel_path = os.path.join(excel_dir, excel_file)

        # Write the DataFrame to the Excel file
        df.to_excel(excel_path, index=False)

        # Autofit the columns in the Excel sheet
        autofit_columns(excel_path)

        # Add table to the Excel sheet
        add_table(excel_path, df)


def autofit_columns(excel_path):
    # Load the Excel workbook
    workbook = load_workbook(excel_path)

    # Get the active sheet
    sheet = workbook.active

    # Autofit columns
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length

    # Save the modified workbook
    workbook.save(excel_path)


def add_table(excel_path, df):
    # Load the Excel workbook
    workbook = load_workbook(excel_path)

    # Get the active sheet
    sheet = workbook.active

    # Convert DataFrame to rows
    rows = dataframe_to_rows(df, index=False, header=True)

    # Determine the range of the table
    start_cell = sheet.cell(row=1, column=1)
    end_cell = sheet.cell(row=df.shape[0]+1, column=df.shape[1])

    # Add the table to the sheet
    table = Table(displayName="Table1", ref=f"{start_cell.coordinate}:{end_cell.coordinate}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Save the modified workbook
    workbook.save(excel_path)

def check_lock():
    return os.path.exists(LOCK_FILE)

def acquire_lock():
    with open(LOCK_FILE, "w") as f:
        f.write("Locked")

def release_lock():
    os.remove(LOCK_FILE)


def main():


    if check_lock():
        print("Script is currently in use by someone else.")
        return

    acquire_lock()

    try:
        csv_dir = 'XXX'
        xlsx_dir = 'XXX'
        convert_csv_to_excel(csv_dir, xlsx_dir)  

    finally:
        release_lock()


if __name__ == "__main__":
    main()





